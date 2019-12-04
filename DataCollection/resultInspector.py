import operator
import os
import pickle
import openpyxl
import nltk

from DataCollection.dataCollection import combine_counts_all_papers, combine_counts_alternate_names
from DataCollection.input_data import get_viruses_data, PUNCTUATION
from helper import print_debug_gene
from DataCollection.paperSelection import open_xml_paper


def get_info_by_gene(virus_data, gene, print_paper_stripping_steps=False):
    index, sorted_index = pickle.load(open(virus_data["counted_file"], "rb"))
    debug_index = pickle.load(open(virus_data["debug_file"], "rb"))
    print(f'Debug info for gene: {gene}')
    for paper, kws in index.items():
        if gene in kws:
            print_debug_gene(paper, debug_index[paper][gene])
            if print_paper_stripping_steps:
                # Open file and make a lowercase list without punctuation and whitespace
                file = open_xml_paper(os.path.join(virus_data['papers_directory'], paper))
                print(file)
                stop_words = set(nltk.corpus.stopwords.words('english'))

                words = file.lower().translate(
                    str.maketrans('\n\t' + PUNCTUATION, ' ' * (len(PUNCTUATION) + 2))).split()
                print(words)
                content = [word for word in words if word not in stop_words and not word.isdigit()]
    print('----\n')


def hsv1_inspection():
    viruses_data = get_viruses_data()
    get_info_by_gene(viruses_data[0], 'ul1')  # toeval
    get_info_by_gene(viruses_data[0], 'ul12')  # toeval
    get_info_by_gene(viruses_data[0], 'us4')  # toeval

    get_info_by_gene(viruses_data[0], 'ul17')  # toeval
    get_info_by_gene(viruses_data[0], 'us9')  # toeval

    get_info_by_gene(viruses_data[0], 'rl1')  # toeval

    get_info_by_gene(viruses_data[0], 'ul4')  # toeval

    get_info_by_gene(viruses_data[0], 'ul14')  # niet duidelijk, waarschijnlijk fout
    get_info_by_gene(viruses_data[0], 'us6')  # correct, ookal is score nipt
    get_info_by_gene(viruses_data[0], 'ul49')  # onduidelijk
    get_info_by_gene(viruses_data[0], 'ul41')  # toeval

    get_info_by_gene(viruses_data[0], 'ul6')  # correct
    get_info_by_gene(viruses_data[0], 'ul46')  # correct
    get_info_by_gene(viruses_data[0], 'ul26')  # correct
    get_info_by_gene(viruses_data[0], 'ul13')  # correct

    get_info_by_gene(viruses_data[0],
                     'ul56')  # Heel nipte score, slechts 1 paper zegt duidelijk dat hij late is, late heeft ook de hoogste score
    get_info_by_gene(viruses_data[0],
                     'ul15')  # Duidelijke score maar afkomstig uit slechts 1 paper, wel correct resultaat
    get_info_by_gene(viruses_data[0],
                     'ul16')  # Duidelijke score maar afkomstig uit slechts 1 paper, wel correct resultaat


def hsv2_inspection():
    viruses_data = get_viruses_data()

    gene_lst = [
        'ul51',  # (0.33): onduidelijk
        'ul18',  # (0.5): klopt: Early-Late (volgens 1 paper)
        'us8',  # (0.5): fout (is early-late ipv late)
        'ul26',  # (0.5): onduidelijk
        'rl1',  # (0.91): fout (IE volgens 1 paper, onduidelijk in andere papers, hierdoor Late hoogste score)
        'ul39',  # (1.2): correct
        'ul50',  # (1.2): correct
        'ul53',  # (1.25): fout (door voorkomen in tabellen)
        'ul31',  # (1.29): toeval
        'ul14',  # (1.57): correct
        'us6',  # (1.71): correct
        'ul42',  # (1.92): correct
        'us3',  # (2.33): papers zijn het niet eens
        'ul23',
        # (2.33): fout door opsommingen, bv: early (E) ul23, late (L) gC, misschien ook afkortingen van early en late rekening houden?
        'ul46',  # (2.39): correct
        'ul36',  # (2.62): correct
        'ul56',
        # (2.94 en 2.9): bijna gelijke score voor early en late, is late volgens 1 paper maar krijgt ook hoge score voor early omdat hij in een andere paper heel vaak voorkomt, toevallig met early
        'ul15',  # (3.54): correct
    ]
    for gene in gene_lst:
        get_info_by_gene(viruses_data[1], gene)


def vzv_inspection():
    viruses_data = get_viruses_data()

    gene_lst = [
        'orf22',  # (0.1): toevallig correct
        'orf37',  # (0.1): toeval
        'orf58',  # (0.25): toeval
        'orf9',  # (0.33): correct
        'orf36',  # (0.36): gaat over ander virus
        'orf31',  # (0.39): correct
        'orf6',  # (0.49): correct
        'orf18',  # (1): fout door naast IE te staan in 1 paper
        'orf1',  # (1.27): IE volgens 1 paper, E volgens een andere
        'mcp',  # (1.62): correct
        'orf10',
        # (1.69): fout, gebruik van afkorting (L) voor echte phase, en heel vaak in combinatie met IE genes gebruikt
        'orf65',  # (2.2): correct
    ]
    for gene in gene_lst:
        get_info_by_gene(viruses_data[2], gene)


def ebv_inspection():
    viruses_data = get_viruses_data()

    gene_lst = [
        'rpms1',  # (0.1): toeval
        'lf1',  # (0.21): toeval
        'lf3',  # (0.6): toeval
        'bbrf2',  # (0.65): correct
        'ebna3',  # (0.79): correct
        'bkrf2',  # (1.03): early volgens 1 paper, late volgens 2 andere
        'bbrf1',  # (1.12): correct
        'bilf2',  # (1.7): correct
    ]
    for gene in gene_lst:
        get_info_by_gene(viruses_data[3], gene)


def hcmv_inspection():
    viruses_data = get_viruses_data()

    gene_lst = [
        'ul21',  # (0.1): toeval
        'ul14',  # (0.2): toeval
        'rl1',  # (0.25): toeval
        'ul11',  # (0.33): onduidelijk
        'ul132',  # (0.5): toeval
        'ul146',  # (0.62): correct
        'ul128',  # (0.62): toeval
        'ul35',  # (0.64): onduidelijk
        'ul142',  # (0.97): onduidelijk
        'ul77',  # (1.04): onduidelijk
        'ul79',  # (1.18): correct
        'ul95',  # (1.27): onduidelijk
        'ul69',  # (1.74): onduidelijk
        'ul144',  # (2.03): onduidelijk
        'ul82',  # (2.15): onduidelijk
        'us27',  # (2.28): correct
        'ul49',  # (2.39): waarschijnlijk correct
        'ul27',  # (2.49): correct
        'ul38',  # (2.54): volgens 2 papers late, volgens 2 papers early
        'ul42',  # (2.59): correct
        'us2',  # (2.96 en 2.95): fout, is early (score 2.95), maar scores verschillen slechts 0.29%
    ]
    for gene in gene_lst:
        get_info_by_gene(viruses_data[4], gene)


def check_with_manual_xlsx():
    manual_xlsx = openpyxl.load_workbook("DataCollection/Data/herpes_lifecycle.xlsx")
    out_xlsx = openpyxl.Workbook()

    viruses_data = get_viruses_data()
    viruses = {'HHV-1': (viruses_data[0], 1, 2),
               'Epstein-barr': (viruses_data[3], 1, 2),
               'VZV': (viruses_data[2], 11, 1)}
    for virus, (virus_data, time_col, id_col) in viruses.items():
        in_sheet = manual_xlsx[virus]
        out_sheet = out_xlsx.create_sheet(virus)
        out_sheet.append(
            ('Protein', 'Calculated time', 'Manual time', 'Correct?', 'IE-score', 'E-score', 'L-score', 'Reason'))

        combined_counts, paper_counts = combine_counts_all_papers(virus_data["counted_file"])
        combined_counts_an, paper_counts_an = combine_counts_alternate_names(combined_counts, paper_counts,
                                                                             virus_data["keywords_file"])
        for row in in_sheet.iter_rows(min_row=4, max_row=200, max_col=12):
            if row[time_col].value is not None:
                max_phase = 'NOT FOUND'
                result = 'incorrect'
                phases_ = {}
                for kws, phases in combined_counts_an.items():
                    kws_lst = kws.split('_')
                    if row[id_col].value.lower() in kws_lst:
                        max_phase = max(phases.items(), key=operator.itemgetter(1))[0]
                        phases_ = phases
                        if max_phase == 'late' and row[time_col].value == 'late':
                            result = 'correct'
                        elif max_phase == 'early' and row[time_col].value == 'early':
                            result = 'correct'
                        elif max_phase == 'immediate-early' and row[time_col].value == 'IE':
                            result = 'correct'

                out_sheet.append(
                    (row[id_col].value, max_phase, row[time_col].value, result, phases_.get('immediate-early', 0),
                     phases_.get('early', 0), phases_.get('late', 0)))

    out_xlsx.remove_sheet(out_xlsx['Sheet'])
    out_xlsx.save('DataCollection/Data/verification.xlsx')


def main():
    check_with_manual_xlsx()


if __name__ == "__main__":
    main()
